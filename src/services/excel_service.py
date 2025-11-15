# -*- coding: utf-8 -*-
"""
Excel generation service for transaction reports
"""
import boto3
import pandas as pd
import io
import logging
from typing import Dict, Any, List
from datetime import datetime
from src.config import Config

logger = logging.getLogger(__name__)

class ExcelService:
    """Service for generating Excel reports"""
    
    def __init__(self):
        self.s3_client = boto3.client('s3', region_name=Config.AWS_REGION)
        self.bucket_name = Config.S3_BUCKET_NAME
    
    def generate_trends_excel(self, trends_data: Dict[str, Any], query_params: Dict[str, Any], phone_number: str = None, transactions: List[Dict[str, Any]] = None) -> tuple:
        """Generate Excel file with trends analysis"""
        try:
            if not trends_data.get('has_data'):
                return None, None
            
            # Create Excel workbook with multiple sheets
            excel_buffer = io.BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # Sheet 1: Executive Summary
                self._create_trends_summary_sheet(writer, trends_data, query_params, phone_number)
                
                # Sheet 2: Product Trends Details
                self._create_trends_details_sheet(writer, trends_data)
                
                # Sheet 3: Weekly Time Series
                self._create_weekly_timeseries_sheet(writer, trends_data)
                
                # Sheet 4: Top Performers
                self._create_top_performers_sheet(writer, trends_data)
                
                # Sheet 5: Heatmap Calendar
                if transactions:
                    self._create_heatmap_calendar_sheet(writer, trends_data, transactions)
                
                # Sheet 6: Insights & Recommendations
                self._create_insights_sheet(writer, trends_data)
            
            excel_buffer.seek(0)
            
            # Upload to S3 with Lima timezone in filename
            from pytz import timezone
            lima_tz = timezone('America/Lima')
            lima_time = datetime.now(lima_tz)
            timestamp = lima_time.strftime('%Y%m%d_%H%M')
            s3_key = f"{Config.S3_IMAGES_PREFIX}analisis_tendencias_{timestamp}_stockai.xlsx"
            
            self.s3_client.put_object(
                Bucket=self.bucket_name,
                Key=s3_key,
                Body=excel_buffer.getvalue(),
                ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ContentDisposition='attachment; filename="analisis_tendencias.xlsx"'
            )
            
            # Generate presigned URL (valid for 24 hours)
            presigned_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': self.bucket_name, 'Key': s3_key},
                ExpiresIn=24 * 3600
            )
            
            # Extract filename from s3_key
            filename = s3_key.split('/')[-1]
            
            logger.info(f"Generated trends Excel report: {presigned_url}")
            return presigned_url, filename
            
        except Exception as e:
            logger.error(f"Error generating trends Excel report: {str(e)}")
            return None, None
    
    def generate_report_excel(self, summary: Dict[str, Any], query_params: Dict[str, Any], phone_number: str = None) -> tuple:
        """Generate Excel report and upload to S3"""
        try:
            products = summary.get('products', [])
            if not products:
                return None
            
            # Create Excel workbook with multiple sheets
            excel_buffer = io.BytesIO()
            
            with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                # Sheet 1: Summary Report
                self._create_summary_sheet(writer, summary, query_params, phone_number)
                
                # Sheet 2: Detailed Data
                self._create_detailed_sheet(writer, products)
                
                # Sheet 3: Charts Data (for manual chart creation)
                self._create_charts_data_sheet(writer, products)
            
            excel_buffer.seek(0)
            
            # Upload to S3 with Lima timezone in filename
            from pytz import timezone
            lima_tz = timezone('America/Lima')
            lima_time = datetime.now(lima_tz)
            timestamp = lima_time.strftime('%Y%m%d_%H%M')
            s3_key = f"{Config.S3_IMAGES_PREFIX}reporte_transacciones_{timestamp}_stockai.xlsx"
            
            self.s3_client.put_object(
                Bucket=self.bucket_name,
                Key=s3_key,
                Body=excel_buffer.getvalue(),
                ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ContentDisposition='attachment; filename="reporte_transacciones.xlsx"'
            )
            
            # Generate presigned URL (valid for 24 hours)
            presigned_url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={'Bucket': self.bucket_name, 'Key': s3_key},
                ExpiresIn=24 * 3600
            )
            
            # Extract filename from s3_key
            filename = s3_key.split('/')[-1]
            
            logger.info(f"Generated Excel report: {presigned_url}")
            return presigned_url, filename
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {str(e)}")
            return None, None
    
    def _create_summary_sheet(self, writer: pd.ExcelWriter, summary: Dict[str, Any], query_params: Dict[str, Any], phone_number: str = None):
        """Create summary sheet with report information"""
        try:
            # Prepare summary data
            summary_data = []
            
            # Header information
            transaction_type_text = ""
            if query_params.get('transaction_type') == 1:
                transaction_type_text = "Ventas"
            elif query_params.get('transaction_type') == 0:
                transaction_type_text = "Compras"
            else:
                transaction_type_text = "Transacciones"
            
            summary_data.append(['Tipo de Reporte', transaction_type_text])
            
            # Group information
            if phone_number and Config.ENABLE_USER_GROUPS:
                from src.models import UserGroupRepository
                user_group_repo = UserGroupRepository()
                user_group = user_group_repo.get_user_group(phone_number)
                
                if user_group and user_group.is_active and user_group.grouped_phone_numbers:
                    member_count = user_group.get_member_count()
                    if user_group.group_name:
                        summary_data.append(['Grupo', f"{user_group.group_name} ({member_count} usuarios)"])
                    else:
                        summary_data.append(['Grupo', f"{member_count} usuarios incluidos"])
            
            # Date range
            if query_params.get('date_from') and query_params.get('date_to'):
                summary_data.append(['Período', f"{query_params['date_from']} al {query_params['date_to']}"])
            elif query_params.get('date_from'):
                summary_data.append(['Desde', query_params['date_from']])
            elif query_params.get('date_to'):
                summary_data.append(['Hasta', query_params['date_to']])
            
            # Products filter
            if query_params.get('products'):
                products_list = ", ".join(query_params['products'])
                summary_data.append(['Productos Filtrados', products_list])
            
            # Totals
            total_cost = float(summary['total_cost'])
            total_transactions = summary['total_transactions']
            total_products = len(summary['products'])
            
            # Get Lima timezone for generation timestamp
            from pytz import timezone
            lima_tz = timezone('America/Lima')
            lima_time = datetime.now(lima_tz)
            generation_time = lima_time.strftime('%Y-%m-%d %H:%M:%S')
            
            summary_data.extend([
                ['', ''],  # Empty row
                ['TOTALES', ''],
                ['Costo Total', f"{total_cost:.2f} PEN"],
                ['Total Transacciones', total_transactions],
                ['Total Productos', total_products],
                ['Fecha Generación', f"{generation_time} (Lima, UTC-5)"]
            ])
            
            # Create DataFrame and write to Excel
            summary_df = pd.DataFrame(summary_data, columns=['Campo', 'Valor'])
            summary_df.to_excel(writer, sheet_name='Resumen', index=False)
            
            # Format the summary sheet
            workbook = writer.book
            worksheet = writer.sheets['Resumen']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Error creating summary sheet: {str(e)}")
    
    def _create_detailed_sheet(self, writer: pd.ExcelWriter, products: List[Dict[str, Any]]):
        """Create detailed products sheet"""
        try:
            # Prepare detailed data
            detailed_data = []
            
            for product_data in products:
                detailed_data.append({
                    'Producto': product_data['product'].title(),
                    'Cantidad Total': float(product_data['total_quantity']),
                    'Unidades': product_data['quantity_units'],
                    'Costo Total (PEN)': float(product_data['total_cost']),
                    'Número de Transacciones': product_data['transaction_count'],
                    'Costo Promedio por Transacción': float(product_data['total_cost']) / product_data['transaction_count'],
                    'Cantidad Promedio por Transacción': float(product_data['total_quantity']) / product_data['transaction_count']
                })
            
            # Create DataFrame
            detailed_df = pd.DataFrame(detailed_data)
            
            # Sort by total cost (descending)
            detailed_df = detailed_df.sort_values('Costo Total (PEN)', ascending=False)
            
            # Write to Excel
            detailed_df.to_excel(writer, sheet_name='Detalle por Producto', index=False)
            
            # Format the detailed sheet
            worksheet = writer.sheets['Detalle por Producto']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Format currency columns
            from openpyxl.styles import NamedStyle
            currency_style = NamedStyle(name='currency', number_format='#,##0.00')
            
            for row in range(2, len(detailed_data) + 2):  # Skip header
                worksheet[f'D{row}'].style = currency_style  # Costo Total
                worksheet[f'F{row}'].style = currency_style  # Costo Promedio
            
        except Exception as e:
            logger.error(f"Error creating detailed sheet: {str(e)}")
    
    def _create_charts_data_sheet(self, writer: pd.ExcelWriter, products: List[Dict[str, Any]]):
        """Create charts data sheet for visualization"""
        try:
            # Top 10 products by cost
            top_products = sorted(products, key=lambda x: float(x['total_cost']), reverse=True)[:10]
            
            chart_data = []
            for product_data in top_products:
                chart_data.append({
                    'Producto': product_data['product'].title(),
                    'Costo Total': float(product_data['total_cost']),
                    'Cantidad Total': float(product_data['total_quantity']),
                    'Transacciones': product_data['transaction_count']
                })
            
            # Create DataFrame
            chart_df = pd.DataFrame(chart_data)
            
            # Write to Excel
            chart_df.to_excel(writer, sheet_name='Top 10 Productos', index=False)
            
            # Format the charts sheet
            worksheet = writer.sheets['Top 10 Productos']
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Error creating charts data sheet: {str(e)}")

    def _create_trends_summary_sheet(self, writer: pd.ExcelWriter, trends_data: Dict[str, Any], query_params: Dict[str, Any], phone_number: str = None):
        """Create trends summary sheet"""
        try:
            summary_data = []
            
            # Header
            transaction_type_text = ""
            if query_params.get('transaction_type') == 1:
                transaction_type_text = "Ventas"
            elif query_params.get('transaction_type') == 0:
                transaction_type_text = "Compras"
            else:
                transaction_type_text = "Transacciones"
            
            summary_data.append(['Tipo de Análisis', f'Tendencias de {transaction_type_text}'])
            
            # Group information
            if phone_number and Config.ENABLE_USER_GROUPS:
                from src.models import UserGroupRepository
                user_group_repo = UserGroupRepository()
                user_group = user_group_repo.get_user_group(phone_number)
                
                if user_group and user_group.is_active and user_group.grouped_phone_numbers:
                    member_count = user_group.get_member_count()
                    if user_group.group_name:
                        summary_data.append(['Grupo', f"{user_group.group_name} ({member_count} usuarios)"])
                    else:
                        summary_data.append(['Grupo', f"{member_count} usuarios incluidos"])
            
            # Date range
            if query_params.get('date_from') and query_params.get('date_to'):
                summary_data.append(['Período', f"{query_params['date_from']} al {query_params['date_to']}"])
                summary_data.append(['Días Analizados', trends_data.get('period_days', 0)])
            
            # Products filter
            if query_params.get('products'):
                products_list = ", ".join(query_params['products'])
                summary_data.append(['Productos Filtrados', products_list])
            
            # Summary stats
            summary_data.extend([
                ['', ''],
                ['RESUMEN', ''],
                ['Total Productos Analizados', trends_data.get('total_products', 0)],
                ['', '']
            ])
            
            # Top growing
            top_growing = trends_data.get('top_growing', [])
            if top_growing:
                summary_data.append(['TOP PRODUCTOS EN CRECIMIENTO', ''])
                for i, product in enumerate(top_growing[:5], 1):
                    summary_data.append([
                        f"{i}. {product['product'].title()}",
                        f"{product['cost_growth_rate']:.1f}% crecimiento semanal"
                    ])
                summary_data.append(['', ''])
            
            # Top declining
            top_declining = trends_data.get('top_declining', [])
            if top_declining:
                summary_data.append(['TOP PRODUCTOS EN DECLIVE', ''])
                for i, product in enumerate(top_declining[:5], 1):
                    summary_data.append([
                        f"{i}. {product['product'].title()}",
                        f"{product['cost_growth_rate']:.1f}% cambio semanal"
                    ])
                summary_data.append(['', ''])
            
            # Generation time
            from pytz import timezone
            lima_tz = timezone('America/Lima')
            lima_time = datetime.now(lima_tz)
            generation_time = lima_time.strftime('%Y-%m-%d %H:%M:%S')
            summary_data.append(['Fecha Generación', f"{generation_time} (Lima, UTC-5)"])
            
            # Create DataFrame
            summary_df = pd.DataFrame(summary_data, columns=['Campo', 'Valor'])
            summary_df.to_excel(writer, sheet_name='Resumen Ejecutivo', index=False)
            
            # Format
            worksheet = writer.sheets['Resumen Ejecutivo']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Error creating trends summary sheet: {str(e)}")
    
    def _create_trends_details_sheet(self, writer: pd.ExcelWriter, trends_data: Dict[str, Any]):
        """Create detailed trends sheet"""
        try:
            product_trends = trends_data.get('product_trends', {})
            
            details_data = []
            for product, trend in product_trends.items():
                if trend.get('trend') in ['error', 'insufficient_data']:
                    continue
                
                details_data.append({
                    'Producto': product.title(),
                    'Tendencia': trend['trend'].upper(),
                    'Semanas Analizadas': trend['weeks_count'],
                    'Crecimiento Semanal (%)': round(trend['cost_growth_rate'], 2),
                    'Cambio Reciente (%)': round(trend.get('recent_change_percent', 0), 2),
                    'Costo Total (PEN)': round(trend['total_cost'], 2),
                    'Cantidad Total': round(trend['total_quantity'], 2),
                    'Promedio Semanal (PEN)': round(trend['avg_weekly_cost'], 2),
                    'Volatilidad': round(trend.get('cost_volatility', 0), 2),
                    'Primera Semana (PEN)': round(trend['first_week_cost'], 2),
                    'Última Semana (PEN)': round(trend['last_week_cost'], 2)
                })
            
            if not details_data:
                return
            
            # Create DataFrame
            details_df = pd.DataFrame(details_data)
            details_df = details_df.sort_values('Crecimiento Semanal (%)', ascending=False)
            details_df.to_excel(writer, sheet_name='Detalle de Tendencias', index=False)
            
            # Format
            worksheet = writer.sheets['Detalle de Tendencias']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Error creating trends details sheet: {str(e)}")
    
    def _create_weekly_timeseries_sheet(self, writer: pd.ExcelWriter, trends_data: Dict[str, Any]):
        """Create weekly time series sheet"""
        try:
            weekly_data = trends_data.get('weekly_data', {})
            
            if not weekly_data:
                return
            
            # Get all unique weeks
            all_weeks = set()
            for product_weeks in weekly_data.values():
                for week in product_weeks:
                    all_weeks.add(week['week_start'])
            
            sorted_weeks = sorted(list(all_weeks))
            
            # Create time series data
            timeseries_data = []
            for product, weeks in weekly_data.items():
                week_dict = {week['week_start']: week['cost'] for week in weeks}
                
                row = {'Producto': product.title()}
                for week in sorted_weeks:
                    row[week] = round(week_dict.get(week, 0), 2)
                
                timeseries_data.append(row)
            
            if not timeseries_data:
                return
            
            # Create DataFrame
            timeseries_df = pd.DataFrame(timeseries_data)
            timeseries_df.to_excel(writer, sheet_name='Serie Temporal Semanal', index=False)
            
            # Format
            worksheet = writer.sheets['Serie Temporal Semanal']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 15)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Error creating weekly timeseries sheet: {str(e)}")
    
    def _create_top_performers_sheet(self, writer: pd.ExcelWriter, trends_data: Dict[str, Any]):
        """Create top performers comparison sheet"""
        try:
            top_growing = trends_data.get('top_growing', [])
            top_declining = trends_data.get('top_declining', [])
            
            performers_data = []
            
            # Add top growing
            performers_data.append({'Categoría': 'TOP CRECIMIENTO', 'Producto': '', 'Tasa (%)': '', 'Costo Total': ''})
            for i, product in enumerate(top_growing[:5], 1):
                performers_data.append({
                    'Categoría': f'{i}',
                    'Producto': product['product'].title(),
                    'Tasa (%)': round(product['cost_growth_rate'], 2),
                    'Costo Total': round(product['total_cost'], 2)
                })
            
            performers_data.append({'Categoría': '', 'Producto': '', 'Tasa (%)': '', 'Costo Total': ''})
            
            # Add top declining
            performers_data.append({'Categoría': 'TOP DECLIVE', 'Producto': '', 'Tasa (%)': '', 'Costo Total': ''})
            for i, product in enumerate(top_declining[:5], 1):
                performers_data.append({
                    'Categoría': f'{i}',
                    'Producto': product['product'].title(),
                    'Tasa (%)': round(product['cost_growth_rate'], 2),
                    'Costo Total': round(product['total_cost'], 2)
                })
            
            if not performers_data:
                return
            
            # Create DataFrame
            performers_df = pd.DataFrame(performers_data)
            performers_df.to_excel(writer, sheet_name='Top Performers', index=False)
            
            # Format
            worksheet = writer.sheets['Top Performers']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 25)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Error creating top performers sheet: {str(e)}")
    
    def _create_insights_sheet(self, writer: pd.ExcelWriter, trends_data: Dict[str, Any]):
        """Create insights and recommendations sheet"""
        try:
            insights = trends_data.get('insights', [])
            
            insights_data = []
            insights_data.append(['INSIGHTS Y RECOMENDACIONES', ''])
            insights_data.append(['', ''])
            
            for i, insight in enumerate(insights, 1):
                insights_data.append([f'{i}.', insight])
            
            insights_data.extend([
                ['', ''],
                ['CÓMO INTERPRETAR ESTE ANÁLISIS', ''],
                ['', ''],
                ['Tendencia CRECIENDO', 'El producto muestra crecimiento sostenido en ventas/compras'],
                ['Tendencia DECRECIENDO', 'El producto muestra declive en ventas/compras'],
                ['Tendencia ESTABLE', 'El producto mantiene niveles constantes'],
                ['', ''],
                ['Crecimiento Semanal (%)', 'Tasa promedio de cambio semana a semana'],
                ['Cambio Reciente (%)', 'Comparación últimas 4 semanas vs período anterior'],
                ['Volatilidad', 'Variabilidad en los datos (mayor = más inestable)'],
                ['', ''],
                ['RECOMENDACIONES GENERALES', ''],
                ['', ''],
                ['Para productos en crecimiento', 'Asegurar inventario suficiente, considerar aumentar precios'],
                ['Para productos en declive', 'Revisar estrategia, considerar promociones o descontinuar'],
                ['Para productos volátiles', 'Monitorear de cerca, ajustar inventario con precaución']
            ])
            
            # Create DataFrame
            insights_df = pd.DataFrame(insights_data, columns=['Categoría', 'Descripción'])
            insights_df.to_excel(writer, sheet_name='Insights', index=False)
            
            # Format
            worksheet = writer.sheets['Insights']
            worksheet.column_dimensions['A'].width = 30
            worksheet.column_dimensions['B'].width = 70
            
        except Exception as e:
            logger.error(f"Error creating insights sheet: {str(e)}")

    def _create_heatmap_calendar_sheet(self, writer: pd.ExcelWriter, trends_data: Dict[str, Any], transactions: List[Dict[str, Any]]):
        """Create heatmap calendar sheet with daily transaction amounts"""
        try:
            from datetime import datetime, timedelta
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            # Aggregate transactions by date
            daily_totals = {}
            for transaction in transactions:
                try:
                    date_str = transaction.get('date_registry', '')
                    if not date_str:
                        continue
                    
                    date = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                    date_key = date.strftime('%Y-%m-%d')
                    cost = float(transaction.get('cost', 0))
                    
                    if date_key not in daily_totals:
                        daily_totals[date_key] = 0
                    daily_totals[date_key] += cost
                    
                except Exception as e:
                    logger.warning(f"Error processing transaction for heatmap: {str(e)}")
                    continue
            
            if not daily_totals:
                return
            
            # Get date range
            dates = sorted(daily_totals.keys())
            start_date = datetime.strptime(dates[0], '%Y-%m-%d')
            end_date = datetime.strptime(dates[-1], '%Y-%m-%d')
            
            # Adjust to start from Monday
            days_to_monday = start_date.weekday()
            calendar_start = start_date - timedelta(days=days_to_monday)
            
            # Create calendar data
            calendar_data = []
            
            # Header row with day names and date range column
            header_row = ['Semana'] + ['L', 'M', 'X', 'J', 'V', 'S', 'D'] + ['Rango']
            calendar_data.append(header_row)
            
            # Generate weeks
            current_date = calendar_start
            week_num = 1
            
            while current_date <= end_date + timedelta(days=6):
                week_start_date = current_date
                week_row = [f'S{week_num}']
                
                for day in range(7):
                    date_key = current_date.strftime('%Y-%m-%d')
                    
                    # Check if date is in range and has data
                    if current_date < start_date or current_date > end_date:
                        week_row.append('')
                    else:
                        amount = daily_totals.get(date_key, 0)
                        week_row.append(round(amount, 2) if amount > 0 else 0)
                    
                    current_date += timedelta(days=1)
                
                # Add date range for this week (start - end)
                week_end_date = current_date - timedelta(days=1)
                date_range = f"{week_start_date.strftime('%d/%m')} - {week_end_date.strftime('%d/%m')}"
                week_row.append(date_range)
                
                calendar_data.append(week_row)
                week_num += 1
            
            # Create DataFrame
            heatmap_df = pd.DataFrame(calendar_data[1:], columns=calendar_data[0])
            heatmap_df.to_excel(writer, sheet_name='Mapa de Calor', index=False)
            
            # Format the heatmap
            worksheet = writer.sheets['Mapa de Calor']
            
            # Calculate color scale based on values
            all_values = [v for v in daily_totals.values() if v > 0]
            if all_values:
                min_val = min(all_values)
                max_val = max(all_values)
                
                # Define color gradient (light green to dark green)
                def get_color_for_value(value):
                    if value == 0 or value == '':
                        return 'F0F0F0'  # Light gray for no data
                    
                    # Normalize value between 0 and 1
                    if max_val > min_val:
                        normalized = (value - min_val) / (max_val - min_val)
                    else:
                        normalized = 1
                    
                    # Color gradient from light green (E8F5E9) to dark green (1B5E20)
                    # Using RGB interpolation
                    r_start, g_start, b_start = 232, 245, 233  # Light green
                    r_end, g_end, b_end = 27, 94, 32  # Dark green
                    
                    r = int(r_start + (r_end - r_start) * normalized)
                    g = int(g_start + (g_end - g_start) * normalized)
                    b = int(b_start + (b_end - b_start) * normalized)
                    
                    return f'{r:02X}{g:02X}{b:02X}'
                
                # Apply formatting
                thin_border = Border(
                    left=Side(style='thin', color='CCCCCC'),
                    right=Side(style='thin', color='CCCCCC'),
                    top=Side(style='thin', color='CCCCCC'),
                    bottom=Side(style='thin', color='CCCCCC')
                )
                
                # Format header row
                for col in range(1, 10):  # A to I (including Rango column)
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFF', size=11)
                    cell.border = thin_border
                
                # Format data cells
                for row_idx in range(2, len(calendar_data) + 1):
                    # Week label
                    week_cell = worksheet.cell(row=row_idx, column=1)
                    week_cell.font = Font(bold=True, size=10)
                    week_cell.alignment = Alignment(horizontal='center', vertical='center')
                    week_cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                    week_cell.border = thin_border
                    
                    # Day cells
                    for col_idx in range(2, 9):  # B to H (days)
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        
                        if value == '' or value is None:
                            color = 'F0F0F0'
                            cell.font = Font(size=9, color='999999')
                        elif value == 0:
                            color = 'FAFAFA'
                            cell.font = Font(size=9, color='CCCCCC')
                        else:
                            color = get_color_for_value(float(value))
                            # Use white text for dark backgrounds
                            if float(value) > (min_val + (max_val - min_val) * 0.6):
                                cell.font = Font(size=9, color='FFFFFF', bold=True)
                            else:
                                cell.font = Font(size=9, color='000000')
                        
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    
                    # Date range cell (column I)
                    range_cell = worksheet.cell(row=row_idx, column=9)
                    range_cell.font = Font(size=9, color='666666')
                    range_cell.alignment = Alignment(horizontal='center', vertical='center')
                    range_cell.fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                    range_cell.border = thin_border
                
                # Set column widths
                worksheet.column_dimensions['A'].width = 10  # Week column
                for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                    worksheet.column_dimensions[col].width = 12  # Day columns
                worksheet.column_dimensions['I'].width = 15  # Date range column
                
                # Set row heights
                for row in range(1, len(calendar_data) + 1):
                    worksheet.row_dimensions[row].height = 25
                
                # Add legend
                legend_start_row = len(calendar_data) + 3
                
                worksheet.cell(row=legend_start_row, column=1, value='LEYENDA:')
                worksheet.cell(row=legend_start_row, column=1).font = Font(bold=True, size=10)
                
                worksheet.cell(row=legend_start_row + 1, column=1, value='Mínimo:')
                worksheet.cell(row=legend_start_row + 1, column=2, value=f'{min_val:.2f} PEN')
                worksheet.cell(row=legend_start_row + 1, column=2).fill = PatternFill(
                    start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'
                )
                
                worksheet.cell(row=legend_start_row + 2, column=1, value='Máximo:')
                worksheet.cell(row=legend_start_row + 2, column=2, value=f'{max_val:.2f} PEN')
                worksheet.cell(row=legend_start_row + 2, column=2).fill = PatternFill(
                    start_color='1B5E20', end_color='1B5E20', fill_type='solid'
                )
                worksheet.cell(row=legend_start_row + 2, column=2).font = Font(color='FFFFFF', bold=True)
                
                worksheet.cell(row=legend_start_row + 3, column=1, value='Sin datos:')
                worksheet.cell(row=legend_start_row + 3, column=2, value='N/A')
                worksheet.cell(row=legend_start_row + 3, column=2).fill = PatternFill(
                    start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'
                )
                
                # Add date range info
                info_row = legend_start_row + 5
                worksheet.cell(row=info_row, column=1, value='Período:')
                worksheet.cell(row=info_row, column=1).font = Font(bold=True, size=10)
                worksheet.cell(row=info_row, column=2, value=f'{dates[0]} al {dates[-1]}')
                
                worksheet.cell(row=info_row + 1, column=1, value='Total días:')
                worksheet.cell(row=info_row + 1, column=2, value=len(dates))
                
                worksheet.cell(row=info_row + 2, column=1, value='Total monto:')
                worksheet.cell(row=info_row + 2, column=2, value=f'{sum(daily_totals.values()):.2f} PEN')
                worksheet.cell(row=info_row + 2, column=2).font = Font(bold=True)
                
        except Exception as e:
            logger.error(f"Error creating heatmap calendar sheet: {str(e)}")
